# -*- encoding:utf-8 -*-

from openpyxl import load_workbook
from project_path import *

class DoExcel:

    def __init__(self,file_name,sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def get_data(self):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        test_data = []      #以列表形式获取数据

        #获取excel里面的值
        for i in range(2,sheet.max_row+1):
            sub_data = {}     #以字典格式存储数据
            sub_data['case_id'] = sheet.cell(i,1).value
            sub_data['用例名称'] = sheet.cell(i,2).value
            sub_data['url'] = sheet.cell(i,3).value
            sub_data['data'] = sheet.cell(i,4).value
            sub_data['method'] = sheet.cell(i,5).value
            sub_data['except'] = sheet.cell(i,6).value

            test_data.append(sub_data)

        return test_data


    def write_back(self,i,result,TestResult):
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(i,7).value = result
        sheet.cell(i,8).value = TestResult
        wb.save(self.file_name)

if __name__ == '__main__':
    test_data = DoExcel(test_case_path,'yiduiyi').get_data()
    print test_data

